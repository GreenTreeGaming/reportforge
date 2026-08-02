from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from reportforge.models import SalesMetrics


INK = "181817"
MUTED = "6B6B66"
ACCENT = "2457E6"
SURFACE = "FFFFFF"
SUBTLE = "F6F6F3"
BORDER_COLOR = "E5E5DF"

HEADER_FILL = PatternFill("solid", fgColor=INK)
HEADER_FONT = Font(color=SURFACE, bold=True)
TITLE_FONT = Font(size=18, bold=True, color=INK)
SUBTITLE_FONT = Font(size=10, color=MUTED)
THIN_BORDER = Border(bottom=Side(style="thin", color=BORDER_COLOR))


def _sanitize_spreadsheet_text(value: Any) -> Any:
    """Prevent uploaded text from being interpreted as a spreadsheet formula."""
    if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _safe_export(frame: pd.DataFrame) -> pd.DataFrame:
    export = frame.copy()
    for column in export.select_dtypes(include=["object", "string"]).columns:
        export[column] = export[column].map(_sanitize_spreadsheet_text)
    return export


def _style_table_sheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False

    for cell in worksheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN_BORDER
    worksheet.row_dimensions[1].height = 24

    for column_cells in worksheet.columns:
        values = [str(cell.value or "") for cell in column_cells[:250]]
        width = min(max(max(map(len, values), default=0) + 2, 11), 38)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def _write_frame(writer, frame: pd.DataFrame, sheet_name: str) -> None:
    _safe_export(frame).to_excel(writer, sheet_name=sheet_name, index=False)


def create_excel_report(
    clean_data: pd.DataFrame,
    rejected_data: pd.DataFrame,
    metrics: SalesMetrics,
    summaries: dict[str, pd.DataFrame],
    advanced: dict[str, object] | None = None,
    *,
    cost_available: bool = True,
) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        clean_export = clean_data.drop(columns=["_source_row"], errors="ignore").copy()
        if "date" in clean_export.columns:
            clean_export["date"] = pd.to_datetime(clean_export["date"]).dt.date
        _write_frame(writer, clean_export, "Clean Data")

        summary_sheets = (
            ("monthly", "Monthly"),
            ("products", "Products"),
            ("customers", "Customers"),
            ("weekdays", "Weekdays"),
            ("anomalies", "Anomalies"),
        )
        for key, sheet_name in summary_sheets:
            frame = summaries.get(key, pd.DataFrame())
            _write_frame(writer, frame, sheet_name)

        if not rejected_data.empty:
            _write_frame(writer, rejected_data, "Rejected Rows")

        if advanced:
            advanced_sheets = (
                ("decision_queue", "Decision Queue"),
                ("revenue_bridge", "Revenue Bridge"),
                ("customer_movement", "Customer Movement"),
                ("rfm", "Customer Health"),
                ("customer_cadence", "Purchase Cadence"),
                ("cross_sell", "Cross Sell"),
                ("product_momentum", "Product Momentum"),
            )
            for key, sheet_name in advanced_sheets:
                frame = advanced.get(key)
                if isinstance(frame, pd.DataFrame):
                    _write_frame(writer, frame, sheet_name)

        workbook = writer.book
        summary = workbook.create_sheet("Executive Summary", 0)
        summary.sheet_view.showGridLines = False
        summary["A1"] = "ReportForge Business Performance Report"
        summary["A1"].font = TITLE_FONT
        summary["A2"] = "A concise, auditable view of performance, risks, and opportunities."
        summary["A2"].font = SUBTITLE_FONT
        summary.append([])
        summary.append(["Metric", "Value"])

        rows: list[tuple[str, Any]] = [
            ("Total revenue", metrics.total_revenue),
            ("Transactions", metrics.transaction_count),
            ("Average order", metrics.average_order_value),
            ("Median order", metrics.median_order_value),
            ("Unique customers", metrics.unique_customers),
            ("Unique products", metrics.unique_products),
            ("Repeat customer rate", metrics.repeat_customer_rate),
            ("Top customer share", metrics.top_customer_share),
            ("Top product share", metrics.top_product_share),
            ("Anomaly count", metrics.anomaly_count),
            ("Rejected rows", len(rejected_data)),
        ]
        if cost_available:
            rows[1:1] = [
                ("Total cost", metrics.total_cost),
                ("Gross profit", metrics.gross_profit),
                ("Gross margin", metrics.gross_margin),
            ]

        if advanced:
            resilience = advanced.get("resilience")
            if resilience is not None:
                rows.extend(
                    [
                        ("Revenue resilience score", getattr(resilience, "score", None)),
                        ("Resilience confidence", getattr(resilience, "confidence", None)),
                        ("Top five customer share", getattr(resilience, "top_five_customer_share", None)),
                        ("Effective customer count", getattr(resilience, "effective_customer_count", None)),
                    ]
                )

        for row in rows:
            summary.append(row)

        header_row = 4
        for cell in summary[header_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="left")

        summary.column_dimensions["A"].width = 31
        summary.column_dimensions["B"].width = 22
        summary.freeze_panes = "A5"

        currency_labels = {
            "Total revenue",
            "Total cost",
            "Gross profit",
            "Average order",
            "Median order",
        }
        percentage_labels = {
            "Gross margin",
            "Repeat customer rate",
            "Top customer share",
            "Top product share",
            "Resilience confidence",
            "Top five customer share",
        }
        for row_number in range(5, 5 + len(rows)):
            label = summary.cell(row=row_number, column=1).value
            value_cell = summary.cell(row=row_number, column=2)
            if label in currency_labels:
                value_cell.number_format = '$#,##0.00'
            elif label in percentage_labels:
                value_cell.number_format = "0.0%"
            elif label == "Revenue resilience score":
                value_cell.number_format = "0.0"

        for sheet_name in workbook.sheetnames:
            if sheet_name != "Executive Summary":
                _style_table_sheet(workbook[sheet_name])

        monthly_sheet = workbook["Monthly"]
        monthly = summaries.get("monthly", pd.DataFrame())
        if not monthly.empty:
            chart = LineChart()
            chart.title = "Monthly Revenue" if not cost_available else "Monthly Revenue and Profit"
            chart.style = 13
            chart.height = 7
            chart.width = 14
            chart.y_axis.title = "Amount"
            chart.x_axis.title = "Month"
            chart.add_data(
                Reference(
                    monthly_sheet,
                    min_col=2,
                    max_col=2,
                    min_row=1,
                    max_row=len(monthly) + 1,
                ),
                titles_from_data=True,
            )
            if cost_available and "profit" in monthly.columns:
                profit_column = int(monthly.columns.get_loc("profit")) + 1
                chart.add_data(
                    Reference(
                        monthly_sheet,
                        min_col=profit_column,
                        max_col=profit_column,
                        min_row=1,
                        max_row=len(monthly) + 1,
                    ),
                    titles_from_data=True,
                )
            chart.set_categories(
                Reference(
                    monthly_sheet,
                    min_col=1,
                    min_row=2,
                    max_row=len(monthly) + 1,
                )
            )
            summary.add_chart(chart, "D4")

        product_sheet = workbook["Products"]
        products = summaries.get("products", pd.DataFrame())
        count = min(len(products), 10)
        if count:
            chart = BarChart()
            chart.title = "Top Products by Revenue"
            chart.style = 10
            chart.height = 7
            chart.width = 14
            chart.add_data(
                Reference(
                    product_sheet,
                    min_col=2,
                    min_row=1,
                    max_row=count + 1,
                ),
                titles_from_data=True,
            )
            chart.set_categories(
                Reference(
                    product_sheet,
                    min_col=1,
                    min_row=2,
                    max_row=count + 1,
                )
            )
            summary.add_chart(chart, "D20")

    return output.getvalue()
